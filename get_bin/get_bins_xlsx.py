import re
import os
from openpyxl import load_workbook

BIN_PATTERN = re.compile(r"\b\d{12}\b")
bins = []

# папка tools/get_bin
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# корень проекта test_reestrs
BASE_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))

# откуда берём xlsx
input_dir = SCRIPT_DIR

# куда сохраняем результат
output_path = os.path.join(BASE_DIR, "data", "bin_list.txt")

print(f"Ищем XLSX файлы в: {input_dir}")

for root, _, files in os.walk(input_dir):
    for filename in files:
        if not filename.lower().endswith(".xlsx"):
            continue

        filepath = os.path.join(root, filename)

        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is None:
                            continue
                        bins.extend(BIN_PATTERN.findall(str(cell)))

            print(f"✓ {filename}")

        except Exception as e:
            print(f"✗ {filename}: {e}")

bins = sorted(set(bins))

os.makedirs(os.path.dirname(output_path), exist_ok=True)

with open(output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(bins))

print(f"\nНайдено БИН: {len(bins)}")
print(f"Сохранено в: {output_path}")

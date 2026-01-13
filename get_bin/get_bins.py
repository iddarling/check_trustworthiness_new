import pdfplumber
import re
import os
from openpyxl import load_workbook

# === Настройки ===
iin_bin_pattern = re.compile(r"\b\d{12}\b")
bin_list = []

# папка tools/get_bin
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# корень проекта test_reestrs
BASE_DIR = os.path.dirname(SCRIPT_DIR)
# откуда читаем файлы
input_dir = SCRIPT_DIR

# куда сохраняем результат
output_path = os.path.join(BASE_DIR, "data", "bin_list.txt")

print(f"Ищем файлы в: {input_dir}")

# === Проход по всем PDF и XLSX ===
for filename in os.listdir(input_dir):
    filepath = os.path.join(input_dir, filename)

    if filename.lower().endswith(".pdf"):
        try:
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        bin_list.extend(iin_bin_pattern.findall(text))
            print(f"✓ PDF: {filename}")
        except Exception as e:
            print(f"✗ PDF {filename}: {e}")

    elif filename.lower().endswith(".xlsx"):
        try:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            bin_list.extend(iin_bin_pattern.findall(str(cell)))
            print(f"✓ XLSX: {filename}")
        except Exception as e:
            print(f"✗ XLSX {filename}: {e}")

# === Обработка и сохранение ===
bin_list = sorted(set(bin_list))[:50]

os.makedirs(os.path.dirname(output_path), exist_ok=True)

with open(output_path, "w", encoding="utf-8") as f:
    f.write("\n".join(bin_list))

print(f"\nИзвлечено {len(bin_list)} ИИН/БИН (лимит 50)")
print(f"Сохранено в: {output_path}")

# === Удаление обработанных файлов ===
for filename in os.listdir(input_dir):
    if filename.lower().endswith((".pdf", ".xlsx")):
        os.remove(os.path.join(input_dir, filename))
        print(f"🗑 Удалён файл: {filename}")

import re
from pathlib import Path
from openpyxl import load_workbook
import csv

BIN_RE = re.compile(r"\b\d{12}\b")

ROOT = Path("СПИСКИ 2026 ГОД рус")
OUT = "bins_found.csv"

rows = []

for xlsx in ROOT.rglob("*.xlsx"):
    if xlsx.name.startswith("~$"):
        continue

    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
    except Exception:
        continue

    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for bin_ in BIN_RE.findall(str(cell)):
                    rows.append({
                        "bin": bin_,
                    })

    wb.close()

# убираем дубли
rows = list({r["bin"]: r for r in rows}.values())

with open(OUT, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=["bin"])
    w.writeheader()
    w.writerows(rows)

print(f"Найдено БИН: {len(rows)}")
print(f"Файл: {OUT}")

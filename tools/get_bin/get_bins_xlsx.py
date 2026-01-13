import re
import os
from openpyxl import load_workbook

# === Настройки ===
BIN_PATTERN = re.compile(r"\b\d{12}\b")
bins = []

# Папка со скриптом
input_dir = os.path.dirname(os.path.abspath(__file__))

# Файл для сохранения результата (корневая data/ директория)
output_path = os.path.join(os.path.dirname(input_dir), "data", "bin_list.txt")

print(f"Ищем XLSX файлы в: {input_dir}")

# === Проход только по XLSX ===
for filename in os.listdir(input_dir):
    if not filename.lower().endswith(".xlsx"):
        continue

    filepath = os.path.join(input_dir, filename)

    try:
        wb = load_workbook(filepath, data_only=True)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue

                    matches = BIN_PATTERN.findall(str(cell))
                    bins.extend(matches)

        print(f"Извлекли БИН из {filename}")

    except Exception as e:
        print(f"Ошибка при обработке {filename}: {e}")

# === Обработка результатов ===
bins = sorted(set(bins))  # уникальные, без лимита

# === Сохранение ===
with open(output_path, "w", encoding="utf-8") as f:
    for bin_value in bins:
        f.write(bin_value + "\n")

print(f"\nВсего найдено БИН: {len(bins)}")
print(f"Результат сохранён в: {output_path}")

# === Удаление файлов (если нужно) ===
# for filename in os.listdir(input_dir):
#     if filename.lower().endswith(".xlsx"):
#         os.remove(os.path.join(input_dir, filename))
